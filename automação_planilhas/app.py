from openpyxl import load_workbook, Workbook

# Ler dados de qualquer planilha
planilha_vendas = load_workbook('vendas_de_lanches.xlsx')
pagina_vendas = planilha_vendas['Sheet']

for linha in pagina_vendas.iter_rows(values_only=True):
    print(linha)


# Automatizar entrada de dados em planilhas
# Inserir dados de qualquer fonte(Word, Banco de dados, outros sistemas) -> Planilha
planilha_contas = Workbook()
pagina1 = planilha_contas.active

with open('anotações.txt', 'r', encoding='utf-8') as arquivo:
    for linha in arquivo:
        pagina1.append(linha.split(','))
        
planilha_contas.save('contas_a_pagar.xlsx')
        
# (Bônus) -Enviar informações dados por e-mail, whatsapp ou telegram